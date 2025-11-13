from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import date

# --- RUTAS ---
logo_path = "C:/Users/Victoria/Downloads/WhatsApp Image 2025-11-11 at 15.09.00"  # cambia TuUsuario por tu nombre real
output_path = "C:/Users/Victoria/Downloads/COTIZACION_CONSTRUCTORA_COSTANCI_ESTELA.xlsx"

# --- CREAR EXCEL ---
wb = Workbook()
ws = wb.active
ws.title = "Cotización C&E"

# --- ESTILOS ---
bold = Font(bold=True)
center = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# --- LOGO ---
try:
    logo = Image(logo_path)
    logo.width = 130
    logo.height = 130
    ws.add_image(logo, "A1")
except Exception as e:
    print(f"No se pudo agregar el logo: {e}")

# --- CABECERA DE DATOS ---
ws["E2"] = "Cliente:"
ws["F2"] = ""
ws["E3"] = "Fecha Emisión:"
ws["F3"] = date.today().strftime("%d/%m/%Y")
ws["E4"] = "Cotización:"
ws["F4"] = ""

# --- BLOQUE PRINCIPAL ---
ws["A6"] = "Vivienda:"
ws["B6"] = ""
ws["D6"] = "Cotización:"
ws["E6"] = "$"
ws["F6"] = ""

ws["A7"] = "m²:"
ws["B7"] = ""
ws["C7"] = "Localidad:"
ws["D7"] = ""

# --- SISTEMA ---
ws["A9"] = "Sistema:"
ws["A10"] = "☐ Obra gris"
ws["B10"] = "☐ Llave en mano"
ws["C10"] = "☐ ECO Premium"
ws["D10"] = "☐ WF"

# --- ESTILO ---
ws["A12"] = "Estilo:"
ws["A13"] = "☐ Minimalista"
ws["B13"] = "☐ Clásico"
ws["C13"] = "☐ Americano"

# --- PROMOCIONES ---
ws["E9"] = "Promociones vigentes:"
ws["E10"] = ""

# --- AGREGAR ARTÍCULO ---
ws["A16"] = "+ Agregar artículo"
ws["A16"].font = Font(bold=True, color="0000CC")

# --- FORMATO Y BORDES ---
for row in ws["A6":"F7"]:
    for cell in row:
        cell.border = thin_border

for row in ws["A9":"D10"]:
    for cell in row:
        cell.border = thin_border

for row in ws["A12":"C13"]:
    for cell in row:
        cell.border = thin_border

# --- AJUSTAR TAMAÑOS ---
ws.column_dimensions["A"].width = 20
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 18
ws.column_dimensions["E"].width = 22
ws.column_dimensions["F"].width = 18
ws.row_dimensions[6].height = 22
ws.row_dimensions[7].height = 22

# --- GUARDAR ARCHIVO ---
wb.save(output_path)
print(f"✅ Archivo creado correctamente en: {output_path}")